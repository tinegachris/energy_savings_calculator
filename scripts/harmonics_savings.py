import json
import csv
from pathlib import Path
import sys
from typing import Dict, List, Tuple, Optional
import xlsxwriter
import openpyxl
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class CalculateHarmonicSavings:
  def __init__(self, config_file: str):
    self.config_file = config_file
    self.config = self.load_config()
    self.months_list = [
      "January", "February", "March", "April", "May", "June",
      "July", "August", "September", "October", "November", "December"
    ]
    self.harmonic_data = self.load_harmonic_data("Harmonic-Distortion-Month.csv")

  def load_config(self) -> Dict:
    """Load configuration from a JSON file."""
    config_path = Path(self.config_file)
    if not config_path.exists():
      logging.error(f"{self.config_file} does not exist. Exiting.")
      sys.exit()
    logging.info(f"{self.config_file} exists. Loading...")
    with open(config_path) as f:
      config = json.load(f)
      logging.info(f"Loaded {self.config_file}.")
    return config

  def load_harmonic_data(self, filename: str) -> Dict[str, List[Dict[str, float]]]:
    """Load harmonic distortion data from CSV file."""
    year = self.config["year"]
    harmonic_file = Path(f"data/harmonic_data/{year}/{filename}")
    if not harmonic_file.exists():
      logging.warning(f"{filename} does not exist. Skipping.")
      return {}
    with open(harmonic_file, mode='r', newline='') as infile:
      reader = csv.DictReader(infile)
      harmonic_data = {month: [] for month in self.months_list}
      for row in reader:
        month = row["Month"]
        if month in harmonic_data:
          harmonic_data[month].append({
            "timestamp": row["Timestamp"],
            "THD_I": float(row["THD_I"]),
            "THD_V": float(row["THD_V"]),
            "Apparent Power (kVA)": float(row["Apparent Power (kVA)"])
          })
      logging.info(f"Loaded {filename}.")
      return harmonic_data

  def calculate_harmonic_savings(self) -> None:
    """Calculate harmonic savings and save to an XLSX file."""
    year = self.config["year"]
    site = self.config["site"]
    results_dir = Path('results')
    results_dir.mkdir(exist_ok=True)
    file_name = results_dir / f"{year}_{site}_Harmonic_Savings.xlsx"
    workbook = xlsxwriter.Workbook(file_name)
    logging.info(f"Created XLSX file: {file_name}")

    summary_worksheet = workbook.add_worksheet(f"{site}_{year}_Harmonic_Savings_Summary")
    logging.info(f"Created summary worksheet: {site}_{year}_Harmonic_Savings_Summary.")

    for month in self.months_list:
      self._process_month_data(workbook, month)

    self._calculate_year_harmonic_savings(workbook)
    workbook.close()
    logging.info(f"Completed calculating harmonic savings and writing: {file_name}")

  def _process_month_data(self, workbook: xlsxwriter.Workbook, month: str) -> None:
    """Process data for a specific month and write to the workbook."""
    year = self.config["year"]
    site = self.config["site"]
    worksheet = workbook.add_worksheet(f"{site}_{month}_Harmonic_Savings")
    month_data = self.harmonic_data.get(month, [])
    if not month_data:
      logging.warning(f"No data for {month}. Skipping.")
      return

    headers = ["Timestamp", "THD_I", "THD_V", "Apparent Power (kVA)", "Non-Compliant Energy (kVAh)", "Energy Losses (kWh)", "Cost Savings ($)"]
    for col_idx, header in enumerate(headers):
      worksheet.write(0, col_idx, header)

    total_non_compliant_energy = 0
    total_energy_losses = 0
    total_cost_savings = 0

    for row_idx, data in enumerate(month_data, start=1):
      non_compliant_energy, energy_losses, cost_savings = self._calculate_savings(data)
      total_non_compliant_energy += non_compliant_energy
      total_energy_losses += energy_losses
      total_cost_savings += cost_savings

      worksheet.write(row_idx, 0, data["timestamp"])
      worksheet.write(row_idx, 1, data["THD_I"])
      worksheet.write(row_idx, 2, data["THD_V"])
      worksheet.write(row_idx, 3, data["Apparent Power (kVA)"])
      worksheet.write(row_idx, 4, non_compliant_energy)
      worksheet.write(row_idx, 5, energy_losses)
      worksheet.write(row_idx, 6, cost_savings)

    worksheet.write(len(month_data) + 1, 0, "Total Non-Compliant Energy (kVAh)")
    worksheet.write(len(month_data) + 1, 1, total_non_compliant_energy)
    worksheet.write(len(month_data) + 2, 0, "Total Energy Losses (kWh)")
    worksheet.write(len(month_data) + 2, 1, total_energy_losses)
    worksheet.write(len(month_data) + 3, 0, "Total Cost Savings ($)")
    worksheet.write(len(month_data) + 3, 1, total_cost_savings)
    logging.info(f"Processed and written data for {month}.")

  def _calculate_savings(self, data: Dict[str, float]) -> Tuple[float, float, float]:
    """Calculate non-compliant energy, energy losses, and cost savings."""
    thd_i_limit = self.config["harmonic_limits"]["THD_I"]
    thd_v_limit = self.config["harmonic_limits"]["THD_V"]
    non_compliant_energy = 0
    energy_losses = 0
    cost_savings = 0

    if data["THD_I"] > thd_i_limit or data["THD_V"] > thd_v_limit:
      non_compliant_energy = data["Apparent Power (kVA)"] * (5 / 60)  # 5 minutes interval
      energy_losses = self.config["loss_factor"] * non_compliant_energy * (data["THD_I"] / 100)
      cost_savings = energy_losses * self.config["cost_per_kwh"]

    return non_compliant_energy, energy_losses, cost_savings

  def _calculate_year_harmonic_savings(self, workbook: xlsxwriter.Workbook) -> None:
    """Calculate yearly savings and update the summary worksheet."""
    site = self.config["site"]
    year = self.config["year"]
    summary_worksheet = workbook.get_worksheet_by_name(f"{site}_{year}_Harmonic_Savings_Summary")
    logging.info(f"Opened summary worksheet: {site}_{year}_Harmonic_Savings_Summary.")

    self._write_summary_headers(summary_worksheet)
    self._aggregate_yearly_savings(summary_worksheet)

  def _write_summary_headers(self, summary_worksheet: xlsxwriter.Workbook.worksheet_class) -> None:
    """Write headers to the summary worksheet."""
    headers = ["Month", "Total Non-Compliant Energy (kVAh)", "Total Energy Losses (kWh)", "Total Cost Savings ($)"]
    for col_idx, header in enumerate(headers):
      summary_worksheet.write(0, col_idx, header)
    logging.info("Written headers to summary worksheet.")

  def _aggregate_yearly_savings(self, summary_worksheet: xlsxwriter.Workbook.worksheet_class) -> None:
    """Aggregate yearly savings and write to the summary worksheet."""
    totals = {
      "total_non_compliant_energy_year": 0,
      "total_energy_losses_year": 0,
      "total_cost_savings_year": 0
    }

    for row_idx, month in enumerate(self.months_list, start=1):
      self._process_month_summary(summary_worksheet, row_idx, month, totals)

    self._write_yearly_totals(summary_worksheet, totals)

  def _process_month_summary(self, summary_worksheet: xlsxwriter.Workbook.worksheet_class, row_idx: int, month: str, totals: Dict[str, float]) -> None:
    """Process and write the summary for a specific month."""
    year = self.config["year"]
    site = self.config["site"]
    worksheet_path = f"results/{year}_{site}_Harmonic_Savings.xlsx"
    worksheet_name = f"{site}_{month}_Harmonic_Savings"

    try:
      wb = openpyxl.load_workbook(worksheet_path, data_only=True)
      worksheet = wb[worksheet_name]
      month_data = self._read_month_data(worksheet)
      self._write_month_summary(summary_worksheet, row_idx, month, month_data)
      self._update_yearly_totals(totals, month_data)
      logging.info(f"Processed and written summary for {month}.")
    except FileNotFoundError:
      logging.error(f"File {worksheet_path} not found. Skipping {month}.")
    except KeyError:
      logging.error(f"Worksheet {worksheet_name} not found in {worksheet_path}. Skipping {month}.")
    except Exception as e:
      logging.error(f"An error occurred while processing {month}: {e}")

  def _read_month_data(self, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, float]:
    """Read month data from the worksheet."""
    return {
      "total_non_compliant_energy": worksheet.cell(row=worksheet.max_row - 3, column=2).value,
      "total_energy_losses": worksheet.cell(row=worksheet.max_row - 2, column=2).value,
      "total_cost_savings": worksheet.cell(row=worksheet.max_row - 1, column=2).value
    }

  def _write_month_summary(self, summary_worksheet: xlsxwriter.Workbook.worksheet_class, row_idx: int, month: str, month_data: Dict[str, float]) -> None:
    """Write the summary for a specific month."""
    summary_worksheet.write(row_idx, 0, month)
    summary_worksheet.write(row_idx, 1, month_data["total_non_compliant_energy"])
    summary_worksheet.write(row_idx, 2, month_data["total_energy_losses"])
    summary_worksheet.write(row_idx, 3, month_data["total_cost_savings"])

  def _update_yearly_totals(self, totals: Dict[str, float], month_data: Dict[str, float]) -> None:
    """Update the yearly totals with the data from a specific month."""
    try:
      totals["total_non_compliant_energy_year"] += float(month_data["total_non_compliant_energy"]) if month_data["total_non_compliant_energy"] is not None else 0
      totals["total_energy_losses_year"] += float(month_data["total_energy_losses"]) if month_data["total_energy_losses"] is not None else 0
      totals["total_cost_savings_year"] += float(month_data["total_cost_savings"]) if month_data["total_cost_savings"] is not None else 0
    except ValueError as e:
      logging.error(f"Value error while updating yearly totals: {e}")
    except TypeError as e:
      logging.error(f"Type error while updating yearly totals: {e}")
    except Exception as e:
      logging.error(f"Unexpected error while updating yearly totals: {e}")

  def _write_yearly_totals(self, summary_worksheet: xlsxwriter.Workbook.worksheet_class, totals: Dict[str, float]) -> None:
    """Write the yearly totals to the summary worksheet."""
    row_idx = len(self.months_list) + 1
    summary_worksheet.write(row_idx, 0, "Yearly Totals")
    summary_worksheet.write(row_idx, 1, totals["total_non_compliant_energy_year"])
    summary_worksheet.write(row_idx, 2, totals["total_energy_losses_year"])
    summary_worksheet.write(row_idx, 3, totals["total_cost_savings_year"])
    logging.info("Written yearly totals to summary worksheet.")

if __name__ == "__main__":
  config_file = "config/harmonic_savings_config.json"
  calculator = CalculateHarmonicSavings(config_file)
  calculator.calculate_harmonic_savings()