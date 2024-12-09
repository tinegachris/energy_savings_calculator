import json
import csv
from pathlib import Path
import sys
from typing import Dict, List, Tuple, Optional
import xlsxwriter
import openpyxl
import logging
from datetime import datetime

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class CalculateGensetSavings:
  def __init__(self, config_file: str):
    self.config_file = config_file
    self.config = self.load_config()
    self.months_list = [
      "January", "February", "March", "April", "May", "June",
      "July", "August", "September", "October", "November", "December"
    ]
    self.solar_yield_data = self.load_yield_data("Solar-Energy-Yield-Month.csv", "Solar_Energy_Yield_Month")
    self.grid_yield_data = self.load_yield_data("Grid-Energy-Yield-Month.csv", "Grid_Energy_Yield_Month")
    self.genset_yield_data = self.load_yield_data("Genset-Energy-Yield-Month.csv", "Genset_Energy_Yield_Month")

  def load_config(self) -> Dict:
    """Load configuration from a JSON file."""
    config_path = Path(self.config_file)
    if not config_path.exists():
      logging.error(f"{self.config_file} does not exist. Exiting.")
      sys.exit()
    logging.info(f"{self.config_file} exists. Loading...")
    with open(config_path) as f:
      config = json.load(f)
      logging.info(f"Loaded {self.config_file}.")
    return config

  def confirm_year_folder(self) -> None:
    """Confirm the existence of the year folder."""
    year_folder = Path(f"data/genset_savings_data/{self.config['year']}")
    if not year_folder.exists():
      logging.error(f"data/genset_savings_data/{self.config['year']} folder does not exist. Exiting.")
      sys.exit()
    logging.info(f"data/genset_savings_data/{self.config['year']} folder exists.")

  def clean_month_csv_files(self) -> None:
    """Remove the 'Conditions Met' column, remove previously calculated savings, and remove empty rows from the CSV file."""
    refined_data = {}
    for month in self.months_list:
      refined_month_data = []
      file_path = Path(f"data/genset_savings_data/{self.config['year']}/{month}-Genset-Savings.csv")
      if not file_path.exists():
        logging.warning(f"data/genset_savings_data/{month}-Genset-Savings.csv does not exist.")
        continue
      with open(file_path, mode='r', newline='') as infile:
        reader = csv.DictReader(infile)
        fieldnames = [field for field in reader.fieldnames if field != "Conditions Met"]
        for row in reader:
          if any(row.values()) and "Total savings" not in row.values():
            refined_row = {field: row[field] for field in fieldnames}
            refined_month_data.append(refined_row)
      refined_data[month] = refined_month_data

    for month, data in refined_data.items():
      file_path = Path(f"data/genset_savings_data/{self.config['year']}/{month}-Genset-Savings.csv")
      with open(file_path, mode='w', newline='') as outfile:
        writer = csv.DictWriter(outfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)
      logging.info(f"Removed 'Conditions Met' column and empty rows from {file_path}.")

  def _parse_time(self, time_str: str) -> datetime:
    """Parse a time string into a datetime object."""
    return datetime.strptime(time_str, "%H:%M:%S")

  def remove_close_time_entries(self) -> None:
    """Filter out rows that are too close in time to the previous row, keeping only those that are spaced by at least 6 minutes from the CSV files."""
    for month in self.months_list:
      file_path = Path(f"data/genset_savings_data/{self.config['year']}/{month}-Genset-Savings.csv")
      if not file_path.exists():
        logging.warning(f"{month}-Genset-Savings.csv does not exist.")
        continue
      filtered_data = []
      with open(file_path, mode='r', newline='') as infile:
        reader = csv.DictReader(infile)
        fieldnames = reader.fieldnames
        prev_time = None
        for row in reader:
          current_time = self._parse_time(row["Time Initiated"])
          if prev_time is not None:
            time_diff = (current_time - prev_time).total_seconds() / 60
            if time_diff >= 6:
              filtered_data.append(row)
          else:
            filtered_data.append(row)
          prev_time = current_time
      with open(file_path, mode='w', newline='') as outfile:
        writer = csv.DictWriter(outfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(filtered_data)
      logging.info(f"Removed rows with 'Time Initiated' - previous row's 'Time Initiated' < 6 minutes from {file_path}.")

  def remove_short_time_entries(self) -> None:
    """Remove rows with 'Time Elapsed (minutes)' less than 1.1 from the CSV files."""
    for month in self.months_list:
      file_path = Path(f"data/genset_savings_data/{self.config['year']}/{month}-Genset-Savings.csv")
      if not file_path.exists():
        logging.warning(f"{month}-Genset-Savings.csv does not exist.")
        continue
      filtered_data = []
      with open(file_path, mode='r', newline='') as infile:
        reader = csv.DictReader(infile)
        fieldnames = reader.fieldnames
        for row in reader:
          if float(row["Time Elapsed (minutes)"]) >= 1.1:
            filtered_data.append(row)
      with open(file_path, mode='w', newline='') as outfile:
        writer = csv.DictWriter(outfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(filtered_data)
      logging.info(f"Removed rows with 'Time Elapsed (minutes)' < 1.1 from {file_path}.")

  def load_yield_data(self, filename: str, yield_column: str) -> Dict[str, float]:
    """Load yield data from CSV file."""
    year = self.config["year"]
    yield_file = Path(f"data/yield_data/{year}/{filename}")
    if not yield_file.exists():
      logging.warning(f"{filename} does not exist. Skipping.")
      return {}
    with open(yield_file, mode='r', newline='') as infile:
      reader = csv.reader(infile)
      headers = next(reader)
      try:
        category_idx = 0
        yield_idx = headers.index(yield_column)
      except ValueError:
        logging.error(f"Required columns not found in {yield_file}. Skipping.")
        return {}
      yield_data = {}
      for row in reader:
        try:
          yield_data[row[category_idx]] = float(row[yield_idx])
        except ValueError:
          logging.warning(f"Invalid data for {row[category_idx]} in {filename}. Skipping.")
      logging.info(f"Loaded {filename}.")
      return yield_data

  def calculate_genset_savings(self) -> None:
    """Copy CSV data of each month into the XLSX file and calculate savings."""
    year = self.config["year"]
    site = self.config["site"]
    results_dir = Path('results')
    results_dir.mkdir(exist_ok=True)
    file_name = results_dir / f"{year}_{site}_Genset_Fuel_Savings.xlsx"
    workbook = xlsxwriter.Workbook(file_name)
    logging.info(f"Created XLSX file: {file_name}")

    summary_worksheet = workbook.add_worksheet(f"{site}_{year}_Savings_Summary")
    logging.info(f"Created summary worksheet: {site}_{year}_Savings_Summary.")

    for month in self.months_list:
      self._process_month_data(workbook, month)

    self._calculate_year_genset_savings(workbook)
    workbook.close()
    logging.info(f"Completed calculating genset savings and writing: {file_name}")

  def _process_month_data(self, workbook: xlsxwriter.Workbook, month: str) -> None:
    """Process data for a specific month and write to the workbook."""
    year = self.config["year"]
    site = self.config["site"]
    worksheet = workbook.add_worksheet(f"{site}_{month}_Savings")
    file_path = Path(f"data/genset_savings_data/{year}/{month}-Genset-Savings.csv")
    if not file_path.exists():
      logging.warning(f"{file_path} does not exist. Skipping.")
      return

    with open(file_path, mode='r', newline='') as infile:
      reader = list(csv.reader(infile))
      for row_idx, row in enumerate(reader):
        for col_idx, cell in enumerate(row):
          self._write_cell(worksheet, row_idx, col_idx, cell)
      logging.info(f"Copied data from {file_path} to {worksheet.name}.")

    self._write_calculations_to_worksheet(worksheet, reader, month)

  def _write_cell(self, worksheet: xlsxwriter.Workbook.worksheet_class, row_idx: int, col_idx: int, cell: str) -> None:
    """Write a cell to the worksheet, handling numbers appropriately."""
    try:
      number = float(cell) if '.' in cell else int(cell)
      worksheet.write_number(row_idx, col_idx, number)
    except ValueError:
      worksheet.write(row_idx, col_idx, cell)

  def _write_calculations_to_worksheet(self, worksheet: xlsxwriter.Workbook.worksheet_class, reader: List[List[str]], month: str) -> None:
    """Write calculations to the worksheet."""
    energy_saving_col_idx = self._get_column_index(reader[0], "Energy Saving (kWh)")
    if energy_saving_col_idx is None:
      logging.warning(f"Energy Saving (kWh) column not found in {month}-Genset-Savings.csv. Skipping calculations.")
      return

    total_kwh_saved, num_outages = self._calculate_savings(reader, energy_saving_col_idx)
    genset_fuel_savings, outage_savings, total_month_genset_savings = self._calculate_costs(total_kwh_saved, num_outages)
    self._write_savings_to_worksheet(worksheet, reader, total_kwh_saved, num_outages, genset_fuel_savings, outage_savings, total_month_genset_savings, month)
    self._adjust_column_widths(worksheet, reader)

  def _get_column_index(self, header: List[str], column_name: str) -> Optional[int]:
    """Get the index of a column in the header."""
    for col_idx, col_name in enumerate(header):
      if col_name == column_name:
        return col_idx
    return None

  def _calculate_savings(self, reader: List[List[str]], energy_saving_col_idx: int) -> Tuple[float, int]:
    """Calculate total kWh saved and number of outages."""
    total_kwh_saved = sum(float(row[energy_saving_col_idx]) for row in reader[1:] if row[energy_saving_col_idx])
    num_outages = len([row for row in reader[1:] if row[energy_saving_col_idx]]) - 1
    return total_kwh_saved, num_outages

  def _calculate_costs(self, total_kwh_saved: float, num_outages: int) -> Tuple[float, float, float]:
    """Calculate genset fuel savings, outage savings, and total month genset savings."""
    cost_fuel_plus_MTCE = self.config["genset_fuel"]["cost_fuel_plus_MTCE"]
    genset_fuel_savings = total_kwh_saved * cost_fuel_plus_MTCE
    cost_per_outage = self.config["genset_fuel"]["cost_per_outage"]
    outage_savings = num_outages * cost_per_outage
    total_month_genset_savings = genset_fuel_savings + outage_savings
    return genset_fuel_savings, outage_savings, total_month_genset_savings

  def _write_savings_to_worksheet(self, worksheet: xlsxwriter.Workbook.worksheet_class, reader: List[List[str]], total_kwh_saved: float, num_outages: int, genset_fuel_savings: float, outage_savings: float, total_month_genset_savings: float, month: str) -> None:
    """Write savings data to the worksheet."""
    worksheet.write(len(reader) + 1, 0, "Total kWh Saved")
    worksheet.write(len(reader) + 1, 1, total_kwh_saved)
    worksheet.write(len(reader) + 3, 0, "Number of Outages")
    worksheet.write(len(reader) + 3, 1, num_outages)
    worksheet.write(len(reader) + 5, 0, "Genset Fuel Savings")
    worksheet.write(len(reader) + 5, 1, genset_fuel_savings)
    worksheet.write(len(reader) + 7, 0, "Outage Savings")
    worksheet.write(len(reader) + 7, 1, outage_savings)
    worksheet.write(len(reader) + 9, 0, "Total Month Genset Savings")
    worksheet.write(len(reader) + 9, 1, total_month_genset_savings)
    worksheet.write(len(reader) + 11, 0, "Solar Yield")
    worksheet.write(len(reader) + 11, 1, self.solar_yield_data.get(month, 0))
    worksheet.write(len(reader) + 13, 0, "Grid Yield")
    worksheet.write(len(reader) + 13, 1, self.grid_yield_data.get(month, 0))
    worksheet.write(len(reader) + 15, 0, "Genset Yield")
    worksheet.write(len(reader) + 15, 1, self.genset_yield_data.get(month, 0))
    logging.info(f"Calculated genset savings for {month}.")

  def _adjust_column_widths(self, worksheet: xlsxwriter.Workbook.worksheet_class, reader: List[List[str]]) -> None:
    """Adjust the column widths based on the content."""
    for col_idx, col_name in enumerate(reader[0]):
      max_len = max(len(str(cell)) for cell in [col_name] + [row[col_idx] for row in reader[1:]])
      worksheet.set_column(col_idx, col_idx, max_len + 2)
    logging.info(f"Adjusted column widths for worksheet {worksheet.name}.")

  def _calculate_year_genset_savings(self, workbook: xlsxwriter.Workbook) -> None:
    """Calculate yearly savings and update the summary worksheet."""
    site = self.config["site"]
    year = self.config["year"]
    summary_worksheet = workbook.get_worksheet_by_name(f"{site}_{year}_Savings_Summary")
    logging.info(f"Opened summary worksheet: {site}_{year}_Savings_Summary.")

    self._write_summary_headers(summary_worksheet)
    self._aggregate_yearly_savings(summary_worksheet)

  def _write_summary_headers(self, summary_worksheet: xlsxwriter.Workbook.worksheet_class) -> None:
    """Write headers to the summary worksheet."""
    headers = ["Month", "Total kWh Saved", "Number of Outages", "Genset Fuel Savings", "Outage Savings", "Total Month Genset Savings", "Solar Yield", "Grid Yield", "Genset Yield"]
    for col_idx, header in enumerate(headers):
      summary_worksheet.write(0, col_idx, header)
    logging.info("Written headers to summary worksheet.")

  def _aggregate_yearly_savings(self, summary_worksheet: xlsxwriter.Workbook.worksheet_class) -> None:
    """Aggregate yearly savings and write to the summary worksheet."""
    totals = {
      "total_kwh_saved_year": 0,
      "total_num_outages_year": 0,
      "total_genset_fuel_savings_year": 0,
      "total_outage_savings_year": 0,
      "total_solar_yield_year": 0,
      "total_grid_yield_year": 0,
      "total_genset_yield_year": 0,
      "total_genset_savings_year": 0
    }

    for row_idx, month in enumerate(self.months_list, start=1):
      self._process_month_summary(summary_worksheet, row_idx, month, totals)

    self._write_yearly_totals(summary_worksheet, totals)

  def _process_month_summary(self, summary_worksheet: xlsxwriter.Workbook.worksheet_class, row_idx: int, month: str, totals: Dict[str, float]) -> None:
    """Process and write the summary for a specific month."""
    year = self.config["year"]
    site = self.config["site"]
    worksheet_path = f"results/{year}_{site}_Genset_Fuel_Savings.xlsx"
    worksheet_name = f"{site}_{month}_Savings"

    try:
      reader = []
      with open(f"data/genset_savings_data/{year}/{month}-Genset-Savings.csv", mode='r', newline='') as infile:
        reader = list(csv.reader(infile))
      wb = openpyxl.load_workbook(worksheet_path, data_only=True)
      worksheet = wb[worksheet_name]
      month_data = self._read_month_data(worksheet, reader)
      self._write_month_summary(summary_worksheet, row_idx, month, month_data)
      self._update_yearly_totals(totals, month_data)
      logging.info(f"Processed and written summary for {month}.")
    except FileNotFoundError:
      logging.error(f"File {worksheet_path} not found. Skipping {month}.")
    except KeyError:
      logging.error(f"Worksheet {worksheet_name} not found in {worksheet_path}. Skipping {month}.")
    except Exception as e:
      logging.error(f"An error occurred while processing {month}: {e}")

  def _read_month_data(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, reader: List[List[str]]) -> Dict[str, float]:
    """Read month data from the worksheet."""
    return {
      "total_kwh_saved": worksheet.cell(row=len(reader) + 2, column=2).value,
      "num_outages": worksheet.cell(row=len(reader) + 4, column=2).value,
      "genset_fuel_savings": worksheet.cell(row=len(reader) + 6, column=2).value,
      "outage_savings": worksheet.cell(row=len(reader) + 8, column=2).value,
      "total_genset_month_savings": worksheet.cell(row=len(reader) + 10, column=2).value,
      "solar_yield": worksheet.cell(row=len(reader) + 12, column=2).value,
      "grid_yield": worksheet.cell(row=len(reader) + 14, column=2).value,
      "genset_yield": worksheet.cell(row=len(reader) + 16, column=2).value
    }

  def _write_month_summary(self, summary_worksheet: xlsxwriter.Workbook.worksheet_class, row_idx: int, month: str, month_data: Dict[str, float]) -> None:
    """Write the summary for a specific month."""
    summary_worksheet.write(row_idx, 0, month)
    summary_worksheet.write(row_idx, 1, month_data["total_kwh_saved"])
    summary_worksheet.write(row_idx, 2, month_data["num_outages"])
    summary_worksheet.write(row_idx, 3, month_data["genset_fuel_savings"])
    summary_worksheet.write(row_idx, 4, month_data["outage_savings"])
    summary_worksheet.write(row_idx, 5, month_data["total_genset_month_savings"])
    summary_worksheet.write(row_idx, 6, month_data["solar_yield"])
    summary_worksheet.write(row_idx, 7, month_data["grid_yield"])
    summary_worksheet.write(row_idx, 8, month_data["genset_yield"])

  def _is_number(self, value: Optional[str]) -> bool:
    """Check if the value is a number."""
    try:
      float(value)
      return True
    except (ValueError, TypeError):
      return False

  def _update_yearly_totals(self, totals: Dict[str, float], month_data: Dict[str, float]) -> None:
    """Update the yearly totals with the data from a specific month."""
    try:
      totals["total_kwh_saved_year"] += float(month_data["total_kwh_saved"]) if self._is_number(month_data["total_kwh_saved"]) else 0
      totals["total_num_outages_year"] += int(month_data["num_outages"]) if self._is_number(month_data["num_outages"]) else 0
      totals["total_genset_fuel_savings_year"] += float(month_data["genset_fuel_savings"]) if self._is_number(month_data["genset_fuel_savings"]) else 0
      totals["total_outage_savings_year"] += float(month_data["outage_savings"]) if self._is_number(month_data["outage_savings"]) else 0
      totals["total_genset_savings_year"] += float(month_data["total_genset_month_savings"]) if self._is_number(month_data["total_genset_month_savings"]) else 0
      totals["total_solar_yield_year"] += float(month_data["solar_yield"]) if self._is_number(month_data["solar_yield"]) else 0
      totals["total_grid_yield_year"] += float(month_data["grid_yield"]) if self._is_number(month_data["grid_yield"]) else 0
      totals["total_genset_yield_year"] += float(month_data["genset_yield"]) if self._is_number(month_data["genset_yield"]) else 0
    except ValueError as e:
      logging.error(f"Value error while updating yearly totals: {e}")
    except TypeError as e:
      logging.error(f"Type error while updating yearly totals: {e}")
    except Exception as e:
      logging.error(f"Unexpected error while updating yearly totals: {e}")

  def _write_yearly_totals(self, summary_worksheet: xlsxwriter.Workbook.worksheet_class, totals: Dict[str, float]) -> None:
    """Write the yearly totals to the summary worksheet."""
    row_idx = len(self.months_list) + 1
    summary_worksheet.write(row_idx, 0, "Yearly Totals")
    summary_worksheet.write(row_idx, 1, totals["total_kwh_saved_year"])
    summary_worksheet.write(row_idx, 2, totals["total_num_outages_year"])
    summary_worksheet.write(row_idx, 3, totals["total_genset_fuel_savings_year"])
    summary_worksheet.write(row_idx, 4, totals["total_outage_savings_year"])
    summary_worksheet.write(row_idx, 5, totals["total_genset_savings_year"])
    summary_worksheet.write(row_idx, 6, totals["total_solar_yield_year"])
    summary_worksheet.write(row_idx, 7, totals["total_grid_yield_year"])
    summary_worksheet.write(row_idx, 8, totals["total_genset_yield_year"])
    logging.info("Written yearly totals to summary worksheet.")

if __name__ == "__main__":
  config_file = "config/savings_config.json"
  calculator = CalculateGensetSavings(config_file)
  calculator.confirm_year_folder()
  calculator.clean_month_csv_files()
  calculator.remove_close_time_entries()
  calculator.remove_short_time_entries()
  calculator.calculate_genset_savings()